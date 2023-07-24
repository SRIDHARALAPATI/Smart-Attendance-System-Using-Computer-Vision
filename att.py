import openpyxl
import tkinter
from tkinter import messagebox


def report():
    wb = openpyxl.load_workbook('attendance.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    total = str(sheet.max_row-1)
    total_present = 0
    for x in range(1, sheet.max_row+1):
        if sheet['B'+str(x)].value == 1:
            total_present+=1
    messagebox.showinfo("Attendance Report", "Total present = " + str(total_present)+"\nTotal Students = "+str(total))
    #print("total present = "+str(total_present))
    #print("total Students = "+total)

    wb.save('attendance.xlsx')